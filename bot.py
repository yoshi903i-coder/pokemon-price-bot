import requests
from bs4 import BeautifulSoup
import re
import subprocess
from datetime import datetime
from openpyxl import load_workbook, Workbook
import os
import time

today_str = datetime.now().strftime("%Y-%m-%d")
now = datetime.now()
sheet_name = str(now.year) + "-" + str(now.month)

all_data = []
page = 1

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.7,en;q=0.3",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
})

# 最初にトップページにアクセスしてクッキーを取得
try:
    session.get("https://www.cardrush-pokemon.jp/", timeout=60)
    time.sleep(2)
except:
    pass

while True:
    url = "https://www.cardrush-pokemon.jp/product-list?page=" + str(page)
    try:
        res = session.get(url, timeout=60)
        soup = BeautifulSoup(res.text, "html.parser")
        items = soup.select(".item_box")
        if not items:
            print("page " + str(page) + " end")
            break
        for item in items:
            name_el = item.select_one(".item_name")
            price_el = item.select_one(".price")
            stock_el = item.select_one(".soldout")
            full_name = name_el.text.strip() if name_el else ""
            state_match = re.search(r'【状態([^】]+)】', full_name)
            state = state_match.group(1) if state_match else ""
            rarity_match = re.search(r'\[([A-Za-z★☆◆\-]+)\]', full_name)
            rarity = rarity_match.group(1) if rarity_match else ""
            clean_name = re.sub(r'【状態[^】]+】', '', full_name)
            clean_name = re.sub(r'\[[^\]]+\]', '', clean_name).strip()
            price_text = price_el.text.strip() if price_el else "0"
            price_num = re.sub(r'[^\d]', '', price_text)
            price = int(price_num) if price_num else 0
            buy_price = int(price * 0.8)
            stock_status = "zaiko_nashi" if stock_el else "zaiko_ari"
            all_data.append({
                "name": clean_name,
                "rarity": rarity,
                "state": state,
                "stock": stock_status,
                "price": price,
                "buy_price": buy_price,
            })
        print("page " + str(page) + " done total=" + str(len(all_data)))
        page += 1
        time.sleep(2)
    except Exception as e:
        print("page " + str(page) + " error: " + str(e))
        page += 1
        time.sleep(5)
    if page > 9999:
        break

print("total: " + str(len(all_data)))

excel_path = "card_prices.xlsx"
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = "card_name"
    ws["B1"] = "rarity"
    ws["C1"] = "state"
    ws["D1"] = "stock"

date_col = None
col = 5
while True:
    val = ws.cell(row=1, column=col).value
    if val == today_str + "_price":
        date_col = col
        break
    elif val is None:
        ws.cell(row=1, column=col).value = today_str + "_price"
        ws.cell(row=1, column=col+1).value = today_str + "_buy"
        ws.cell(row=1, column=col+2).value = today_str + "_ratio"
        ws.cell(row=1, column=col+3).value = today_str + "_diff"
        date_col = col
        break
    col += 4

existing_rows = {}
for row in range(2, ws.max_row + 1):
    key = (ws.cell(row=row, column=1).value, ws.cell(row=row, column=3).value)
    existing_rows[key] = row

next_row = max(ws.max_row + 1, 2)

for d in all_data:
    key = (d["name"], d["state"])
    if key in existing_rows:
        r = existing_rows[key]
    else:
        ws.cell(row=next_row, column=1).value = d["name"]
        ws.cell(row=next_row, column=2).value = d["rarity"]
        ws.cell(row=next_row, column=3).value = d["state"]
        ws.cell(row=next_row, column=4).value = d["stock"]
        r = next_row
        existing_rows[key] = next_row
        next_row += 1
    ws.cell(row=r, column=2).value = d["rarity"]
    ws.cell(row=r, column=4).value = d["stock"]
    ws.cell(row=r, column=date_col).value = d["price"]
    ws.cell(row=r, column=date_col+1).value = d["buy_price"]
    if date_col > 5:
        prev_price = ws.cell(row=r, column=date_col-4).value
        if prev_price and prev_price > 0 and d["price"] > 0:
            ws.cell(row=r, column=date_col+2).value = round(d["price"] / prev_price, 3)
            ws.cell(row=r, column=date_col+3).value = d["price"] - prev_price

if "graph" not in wb.sheetnames:
    wg = wb.create_sheet("graph")
    wg["A1"] = "card_name:"
    wg["B1"] = ""
    wg["A2"] = "start_date:"
    wg["B2"] = ""
    wg["A3"] = "end_date:"
    wg["B3"] = ""
    wg["A5"] = "date"
    wg["B5"] = "price"
    wg["C5"] = "buy_price"

wb.save(excel_path)
print("saved!")

subprocess.run(["git", "config", "user.email", "bot@example.com"])
subprocess.run(["git", "config", "user.name", "PriceBot"])
subprocess.run(["git", "add", "card_prices.xlsx"])
subprocess.run(["git", "commit", "-m", "update " + today_str])
subprocess.run(["git", "push"])
print("done!")
